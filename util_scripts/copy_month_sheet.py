"""
This script copies month sheet to year and (or) total workbook

Arguments:
sheet_name: name of the file with month data (or 'last' for the last one)
if_year: 1 if the worksheet is to be copied to a yearly workbook
if_total: 1 if the worksheet is to be copied to the total workbook
"""

import os
import sys
import xlsxwriter
from copy import copy
from openpyxl import load_workbook

sheet_name = sys.argv[1]
if_year = sys.argv[2]
if_total = sys.argv[3]


# 1. Get paths of month, year and total data
folder_path_file = open("../path.txt", "r")
root_path = folder_path_file.read()
month_folder_path = root_path + "/monthly data/"
year_folder_path = root_path + "/yearly data/"
total_folder_path = root_path + "/total data/"
folder_path_file.close()

# 2. Get paths of desired files
file_name = (os.listdir(month_folder_path)[-1]
             if sheet_name == "last" else sheet_name)

month_sheet_path = month_folder_path + file_name
year_sheet_path = year_folder_path + file_name[:4] + ".xlsx"
total_sheet_path = total_folder_path + "total.xlsx"

# 3. Get first sheet of the month data .xlsx file and list of width
month_workbook = load_workbook(month_sheet_path)
month_worksheet = month_workbook.active
col_widths = []
for col_num in range(0, 20):
    col_letter = xlsxwriter.utility.xl_col_to_name(col_num)
    width = month_worksheet.column_dimensions[col_letter].width
    col_widths.append(width)

# 4. Paste the worksheet to year and total
def copy_cell(source_cell, coord, tgt):
    tgt.cell(coord[0], coord[1]).value = source_cell.value
    tgt.cell(coord[0], coord[1]).comment = source_cell.comment
    if source_cell.has_style:
        tgt.cell(coord[0], coord[1])._style = copy(source_cell._style)
    return tgt.cell(coord[0], coord[1])


# year
if if_year == "1":
    year_workbook = load_workbook(year_sheet_path)
    year_workbook.create_sheet(file_name[5:7])
    year_worksheet = year_workbook[year_workbook.sheetnames[-1]]

    for i in range(1, 100):
        for j in range(1, 100):
            copy_cell(month_worksheet.cell(i, j), [i, j], year_worksheet)

    for colnum in range(0, len(col_widths)):
        colletter = xlsxwriter.utility.xl_col_to_name(colnum)
        year_worksheet.column_dimensions[colletter].width \
            = col_widths[colnum]

    year_workbook.save(year_sheet_path)

# total
if if_total == "1":
    total_workbook = load_workbook(total_sheet_path)
    total_workbook.create_sheet(file_name[5:7]+"."+file_name[2:4])
    total_worksheet = total_workbook[total_workbook.sheetnames[-1]]

    for i in range(1, 100):
        for j in range(1, 100):
            copy_cell(month_worksheet.cell(i, j), [i, j], total_worksheet)

    for colnum in range(0, len(col_widths)):
        colletter = xlsxwriter.utility.xl_col_to_name(colnum)
        total_worksheet.column_dimensions[colletter].width \
            = col_widths[colnum]

    total_workbook.save(total_sheet_path)
