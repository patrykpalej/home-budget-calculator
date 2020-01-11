"""
This script deletes last sheet in year and (or) total workbook

Arguments:
year_nr: two-digits year number
if_year: 1 if a sheet is to be deleted from the yearly workbook
if_total: 1 if a sheet is to be deleted from the total workbook
"""

import sys
from openpyxl import load_workbook

year_nr = sys.argv[1]
if_year = sys.argv[2]
if_total = sys.argv[3]


# 1. Get paths of year and total data
folder_path_file = open("../path.txt", "r")
root_path = folder_path_file.read()
year_folder_path = root_path + "/yearly data/"
total_folder_path = root_path + "/total data/"
folder_path_file.close()

# 2. Delete last sheet in year and total
if if_year == "1":
    year_sheet_path = year_folder_path + "20" + year_nr + ".xlsx"

    year_workbook = load_workbook(year_sheet_path)
    last_sheet = year_workbook.sheetnames[-1]
    year_workbook.remove(year_workbook[last_sheet])

    year_workbook.save(year_sheet_path)

if if_total == "1":
    total_sheet_path = total_folder_path + "total.xlsx"

    total_workbook = load_workbook(total_sheet_path)
    last_sheet = total_workbook.sheetnames[-1]
    total_workbook.remove(total_workbook[last_sheet])

    total_workbook.save(total_sheet_path)
