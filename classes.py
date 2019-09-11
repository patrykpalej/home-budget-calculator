import openpyxl


class MyWorkbook:
    def __init__(self, file_path):
        # Loads excel file and divides it into the sheets
        self.mywb = openpyxl.load_workbook(filename=file_path,
                                           data_only=True)

        self.sheets_list = []  # list of parsed sheets objects
        self.sheets_dict = dict()  # dict of parsed sheets objects
        for ws in self.mywb.sheetnames:
            _myWs = MyWorksheet(ws, self.mywb)
            self.sheets_list.append(_myWs)
            self.sheets_dict[ws] = _myWs

        # -- ! Sums up all sheets of the workbook into one structure ! --
        # 1. balances, spendings' sums, incomes and earnings
        self.balance = [0, 0]
        self.sum_basic = 0
        self.sum_addit = 0
        self.sum_giftdon = 0
        self.savings = 0
        self.sum_total = 0
        self.incomes = 0
        self.earnings = 0
        self.incomes_dict = {}
        self.earnings_dict = {}

        for sheet in self.sheets_list:
            self.balance = [self.balance[0] + sheet.balance[0],
                             self.balance[1] + sheet.balance[1]]
            self.sum_basic += sheet.sum_basic
            self.sum_addit += sheet.sum_addit
            self.sum_giftdon += sheet.sum_giftdon
            self.savings += sheet.savings
            self.sum_total += sheet.sum_total
            self.incomes += sheet.incomes
            self.earnings += sheet.earnings

            for source in sheet.incomes_dict:
                if source in self.incomes_dict:
                    self.incomes_dict[source] += sheet.incomes_dict[source]
                    self.earnings_dict[source] += sheet.earnings_dict[source]
                else:
                    self.incomes_dict[source] = sheet.incomes_dict[source]
                    self.earnings_dict[source] = sheet.earnings_dict[source]

        # 2. Sums of the spendings within categories
        self.cats_sums = {}
        for sheet in self.sheets_list:
            for cat_name in sheet.cats_sums:
                if cat_name in self.cats_sums:
                    self.cats_sums[cat_name] += sheet.cats_sums[cat_name]
                else:
                    self.cats_sums[cat_name] = sheet.cats_sums[cat_name]

        self.cats_sums_list = list(self.cats_sums.values())
        self.cats_names = list(self.cats_sums.keys())

        # 3. Individual spendings with descriptions
        self.spends_values_yr = {}
        self.spends_items_yr = {}
        self.spends_monthlabel_yr = {}

        for s, sheet in enumerate(self.sheets_list):
            for cat_name in sheet.spends_values:
                if cat_name in self.spends_values_yr:
                    self.spends_values_yr[cat_name] = \
                        self.spends_values_yr[cat_name] \
                        + sheet.spends_values[cat_name]

                    self.spends_items_yr[cat_name] = \
                        self.spends_items_yr[cat_name] \
                        + sheet.spends_items[cat_name]

                    self.spends_monthlabel_yr[cat_name] = \
                        self.spends_monthlabel_yr[cat_name] \
                        + [s+1]*len(sheet.spends_items[cat_name])
                else:
                    self.spends_values_yr[cat_name] = \
                        sheet.spends_values[cat_name]

                    self.spends_items_yr[cat_name] = \
                        sheet.spends_items[cat_name]

                    self.spends_monthlabel_yr[cat_name] = \
                        [s+1]*len(sheet.spends_items[cat_name])

# ------------------------------------------------------------------------------


class MyWorksheet:
    def __init__(self, sheet_name, myWb):
        # Sheet itself:
        self.myws = myWb[sheet_name]

        # -- ! Parsing the sheet ! --
        # 1. balances, spendings' sums, incomes and earnings
        self.balance = [self.myws['B1'].value, self.myws['B2'].value]

        self.sum_basic = self.myws['A6'].value
        self.sum_addit = self.myws['B6'].value
        self.sum_giftdon = self.myws['C6'].value
        self.savings = self.myws['D6'].value
        self.sum_total = self.myws['A9'].value

        self.incomes = self.myws['E1'].value
        self.earnings = self.myws['E2'].value
        self.sources = []
        self.incomes_dict = {}
        self.earnings_dict = {}
        col_nr = 6

        while True:
            try:
                len(self.myws.cell(row=3, column=col_nr).value)    # test
                _incomes = self.myws.cell(row=1, column=col_nr).value
                _earnings = self.myws.cell(row=2, column=col_nr).value
                _name = self.myws.cell(row=3, column=col_nr).value

                self.sources.append(_name)
                self.incomes_dict[self.sources[-1]] = _incomes
                self.earnings_dict[self.sources[-1]] = _earnings
            except:
                break

            col_nr += 1

        # 2. Sums of the spendings within categories
        self.cats_names = []
        self.cats_sums = {}
        col_nr = 2

        while True:
            try:
                len(self.myws.cell(row=10, column=col_nr).value)  # test
                _sum = round(self.myws.cell(row=9, column=col_nr).value, 2)
                _name = self.myws.cell(row=10, column=col_nr).value

                self.cats_names.append(_name)
                self.cats_sums[self.cats_names[-1]] = _sum
            except:
                break

            col_nr += 1

        self.cats_sums_list = list(self.cats_sums.values())

        # 3. Individual spendings with descriptions
        self.spends_values = {}
        self.spends_items = {}
        col_nr = 2
        i = 0

        while True:

            try:
                self.spends_values[self.cats_names[i]] = []
                self.spends_items[self.cats_names[i]] = []
            except:
                pass

            row_nr = 11

            if self.myws.cell(row_nr-1, col_nr).value is not None:

                while True:
                    _spending = self.myws.cell(row_nr, col_nr).value
                    if _spending is not None:
                        self.spends_values[self.cats_names[i]].append(_spending)
                        try:
                            _item = self.myws.cell(row_nr, col_nr).comment.text
                            self.spends_items[self.cats_names[i]]\
                                .append(_item[_item.find(':\n')+2:])
                        except:
                            self.spends_items[self.cats_names[i]].append('')
                    else:
                        col_nr += 1
                        i += 1
                        break

                    row_nr += 1
            else:
                break
