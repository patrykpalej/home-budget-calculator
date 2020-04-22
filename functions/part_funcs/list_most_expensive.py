import openpyxl
import numpy as np
from math import floor
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side


def list_most_expensive(results_dir, part_label, my_workbook, start_label):

    wb_to_export = openpyxl.Workbook()

    # Data preparation
    all_spends_values = my_workbook.spends_values_yr
    all_spends_items = my_workbook.spends_items_yr
    all_spends_months = my_workbook.spends_monthlabel_yr

    high_spends_values = []
    high_spends_items = []
    high_spends_dates = []
    high_spends_categories = []

    value_treshold = 100

    mdict = {1: "Styczeń", 2: "Luty", 3: "Marzec", 4: "Kwiecień", 5: "Maj",
             6: "Czerwiec", 7: "Lipiec", 8: "Sierpień", 9: "Wrzesień",
             10: "Październik", 11: "Listopad", 12: "Grudzień", 0: "Grudzień"}

    for values, items, months, cat_name in zip(all_spends_values.values(),
                                               all_spends_items.values(),
                                               all_spends_months.values(),
                                               all_spends_values.keys()):
        for value, item, month in zip(values, items, months):
            if value >= value_treshold and cat_name not in ["Mieszkanie",
                                                            "Jedzenie"]:
                high_spends_values.append(value)
                high_spends_items.append(item)

                month_label = ((month + start_label[0] - 1)
                               if month + start_label[0] - 1 <= 12
                               else (month + start_label[0] - 1) % 12)
                year_label = start_label[1] \
                    + floor((month-1 + start_label[0] - 1) / 12)
                high_spends_dates.append(mdict[month_label] + " 20" +
                                         str(year_label))

                high_spends_categories.append(cat_name)

    sorted_indices = list(np.argsort(high_spends_values))
    sorted_indices.reverse()

    # Excel filling
    # -- main table
    ws = wb_to_export.active
    ws.title = "> 100zł"

    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))

    for i, colname in enumerate(["id", "Nazwa", "Wartość [zł]", "Kategoria",
                                 "Miesiąc"]):
        ws.cell(1, i+1).value = colname
        ws.cell(1, i+1).fill = PatternFill(fgColor="93e1e6", fill_type="solid")

    for i, one_list in enumerate([high_spends_items, high_spends_values,
                                  high_spends_categories, high_spends_dates]):
        for row, index in enumerate(sorted_indices):
            ws.cell(2+row, 1).value = row+1
            ws.cell(2+row, i+2).value = one_list[index]

    for row in range(1, len(high_spends_values)+2):
        for i in range(5):
            ws.cell(row, i+1).alignment = Alignment(horizontal="center")
            ws.cell(row, i+1).border = thin_border

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = max([len(i) for i in
                                           high_spends_items]) + 1
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = max([len(i) for i in
                                           high_spends_categories])
    ws.column_dimensions["E"].width = max([len(i) for i in
                                           high_spends_dates])

    # -- small table
    unique_categories = list(np.unique(high_spends_categories))
    unique_cat_dict = dict()
    for k in unique_categories:
        unique_cat_dict[k] = []

    for unique_cat in unique_categories:
        for val, cat in zip(high_spends_values, high_spends_categories):
            if cat == unique_cat:
                unique_cat_dict[unique_cat].append(val)

    summary_cat_dict = dict()
    for k in unique_cat_dict.keys():
        summary_cat_dict[k] = [len(unique_cat_dict[k]),
                               np.mean(unique_cat_dict[k])]

    ws.cell(2, 8).value = "Kategoria"
    ws.cell(2, 9).value = "Ilość"
    ws.cell(2, 10).value = "Średnia wartość [zł]"

    ws.cell(2, 8).fill = PatternFill(fgColor="93e1e6", fill_type="solid")
    ws.cell(2, 9).fill = PatternFill(fgColor="93e1e6", fill_type="solid")
    ws.cell(2, 10).fill = PatternFill(fgColor="93e1e6", fill_type="solid")

    for row in range(len(summary_cat_dict.keys()) + 1):
        for i in range(3):
            ws.cell(row + 2, i + 8).alignment = Alignment(horizontal="center")
            ws.cell(row + 2, i + 8).border = thin_border

    i = 0
    values_unsorted = [x[0] for x in summary_cat_dict.values()]
    arg_sort = list(np.argsort(np.argsort(values_unsorted)))

    for cat, summary in summary_cat_dict.items():
        ws.cell(len(arg_sort) - arg_sort[i] + 2, 8).value = cat
        ws.cell(len(arg_sort) - arg_sort[i] + 2, 9).value = summary[0]
        ws.cell(len(arg_sort) - arg_sort[i] + 2, 10).value = round(summary[1],
                                                                   2)
        i += 1

    ws.column_dimensions["H"].width = max([len(i) for i in
                                           summary_cat_dict.keys()]) + 1
    ws.column_dimensions["J"].width = 18
    # -----

    wb_to_export.save(results_dir + "/" + part_label
                      + " - największe wydatki.xlsx")
