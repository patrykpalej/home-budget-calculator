import openpyxl
import numpy as np
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side


def create_xlsx_report(results_dir, part_label, my_workbook, start_label,
                       n_of_months):

    wb_to_export = openpyxl.Workbook()

    def export_to_excel(cat_name, num_of_ws):
        values = my_workbook.spends_values_yr[cat_name]
        items = my_workbook.spends_items_yr[cat_name]
        monthlabels = my_workbook.spends_monthlabel_yr[cat_name]

        mdict = {1: "Styczeń", 2: "Luty", 3: "Marzec", 4: "Kwiecień",
                 5: "Maj", 6: "Czerwiec", 7: "Lipiec", 8: "Sierpień",
                 9: "Wrzesień", 10: "Październik", 11: "Listopad",
                 12: "Grudzień"}

        values_dict = dict()
        items_dict = dict()

        for month_num in np.unique(monthlabels):
            _one_month_list = [values[i] for i, j in enumerate(monthlabels)
                               if j == month_num]
            values_dict[month_num] = _one_month_list

            _one_month_list = [items[i] for i, j in enumerate(monthlabels)
                               if j == month_num]
            items_dict[month_num] = _one_month_list

        ws = wb_to_export.create_sheet(cat_name, num_of_ws)

        if len(items):
            ws.column_dimensions["A"].width = max([len(i) for i in items]) + 2

        row = 0
        month = start_label[0]
        year = start_label[1]
        for number in range(1, n_of_months + 1):  # list(values_dict.keys()):
            if number in monthlabels:
                row += 1
                ws.cell(row, 1).value = mdict[month] + "  " + str(month) \
                    + ".20" + str(year)
                ws.merge_cells(start_row=row, start_column=1, end_row=row,
                               end_column=2)
                ws.cell(row, 1).alignment = Alignment(horizontal="center")
                ws.cell(row, 1).fill = PatternFill(fgColor="93e1e6",
                                                   fill_type="solid")

                row += 1
                ws.cell(row, 1).value = "Co?"
                ws.cell(row, 2).value = "Kwota [zł]"
                ws.cell(row, 1).alignment = Alignment(horizontal="center")
                ws.cell(row, 2).alignment = Alignment(horizontal="center")
                ws.cell(row, 1).fill = PatternFill(fgColor="daf6f7",
                                                   fill_type="solid")
                ws.cell(row, 2).fill = PatternFill(fgColor="daf6f7",
                                                   fill_type="solid")

                for i, val in enumerate(values_dict[number]):
                    row += 1
                    ws.cell(row, 1).value = items_dict[number][i]
                    ws.cell(row, 2).value = val

            month += 1
            if month > 12:
                month -= 12
                year += 1

        thin_border = Border(left=Side(style="thin"),
                             right=Side(style="thin"),
                             top=Side(style="thin"),
                             bottom=Side(style="thin"))
        for r in range(row):
            ws.cell(row=r + 1, column=1).border = thin_border
            ws.cell(row=r + 1, column=2).border = thin_border

        return wb_to_export

    wb_to_export.remove_sheet(wb_to_export.active)

    list_of_categories \
        = ["Rzeczy i sprzęty", "Rozrywka", "Transport i noclegi",
           "Abonamenty i usługi"]

    for i, cat in enumerate(list_of_categories):
        wb_to_export = export_to_excel(cat, i + 1)

    wb_to_export.save(results_dir + "/" + part_label
                      + " - zestawienie wydatków.xlsx")
