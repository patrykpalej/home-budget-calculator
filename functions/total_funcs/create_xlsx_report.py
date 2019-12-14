import openpyxl
import numpy as np
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side


def create_xlsx_report(results_dir, total_label, spendings_list, incomes_list,
                       earnings_list, surplus_list, my_workbook, start_label,
                       n_of_months):

    summary_wb = openpyxl.Workbook()

    def export_to_excel(cat_name, num_of_ws):
        # 1. Summary of the period
        # a) first table
        ws = summary_wb.active
        ws.title = "Ogólne"

        ws.cell(1, 2).value = "Wydatki [zł]:"
        ws.cell(1, 3).value = "Zarobki [zł]:"
        ws.cell(1, 4).value = "Przychody [zł]:"
        ws.cell(1, 5).value = "Nadwyżki [zł]:"

        ws.cell(2, 1).value = "Średnia: "
        ws.cell(3, 1).value = "Mediana: "
        ws.cell(4, 1).value = "Std: "

        spendings_mms = [np.mean(spendings_list), np.median(spendings_list),
                         np.std(spendings_list)]
        earnings_mms = [np.mean(earnings_list), np.median(earnings_list),
                        np.std(earnings_list)]
        incomes_mms = [np.mean(incomes_list), np.median(incomes_list),
                       np.std(incomes_list)]
        surplus_mms = [np.mean(surplus_list), np.median(surplus_list),
                       np.std(surplus_list)]

        ws.cell(2, 2).value = round(spendings_mms[0], 2)
        ws.cell(3, 2).value = round(spendings_mms[1], 2)
        ws.cell(4, 2).value = round(spendings_mms[2], 2)

        ws.cell(2, 3).value = round(earnings_mms[0], 2)
        ws.cell(3, 3).value = round(earnings_mms[1], 2)
        ws.cell(4, 3).value = round(earnings_mms[2], 2)

        ws.cell(2, 4).value = round(incomes_mms[0], 2)
        ws.cell(3, 4).value = round(incomes_mms[1], 2)
        ws.cell(4, 4).value = round(incomes_mms[2], 2)

        ws.cell(2, 5).value = round(surplus_mms[0], 2)
        ws.cell(3, 5).value = round(surplus_mms[1], 2)
        ws.cell(4, 5).value = round(surplus_mms[2], 2)

        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["H"].width = 10
        ws.column_dimensions["I"].width = 18

        for r in range(2, 6):
            ws.cell(r, 1).alignment = Alignment(horizontal="right")

        for c in range(2, 6):
            ws.cell(1, c).alignment = Alignment(horizontal="center")

        thin_border = Border(left=Side(style="thin"),
                             right=Side(style="thin"),
                             top=Side(style="thin"),
                             bottom=Side(style="thin"))

        for r in range(4):
            for c in range(5):
                ws.cell(r + 1, c + 1).border = thin_border

        # b) second table
        ws.merge_cells(start_row=1, start_column=8, end_row=1, end_column=9)
        ws.cell(1, 8).value = "Całkowita nadwyżka przychodów:"
        ws.cell(2, 8).value = "[zł]"
        ws.cell(2, 9).value = "[% przychodów]"
        ws.cell(3, 8).value = sum(surplus_list)
        ws.cell(3, 9).value = round(
            sum(surplus_list) / sum(incomes_list) * 100, 2)

        for r in range(1, 4):
            for c in range(8, 10):
                ws.cell(r, c).border = thin_border
                ws.cell(r, c).alignment = Alignment(horizontal="center")

        # 2. Listing the spendings
        values = my_workbook.spends_values_yr[cat_name]
        items = my_workbook.spends_items_yr[cat_name]
        monthlabels = my_workbook.spends_monthlabel_yr[cat_name]

        mdict = {1: "Styczeń", 2: "Luty", 3: "Marzec", 4: "Kwiecień", 5: "Maj",
                 6: "Czerwiec", 7: "Lipiec", 8: "Sierpień", 9: "Wrzesień",
                 10: "Październik", 11: "Listopad", 12: "Grudzień"}

        values_dict = dict()
        items_dict = dict()

        for month_num in np.unique(monthlabels):
            _one_month_list = [values[i] for i, j in enumerate(monthlabels)
                               if j == month_num]
            values_dict[month_num] = _one_month_list

            _one_month_list = [items[i] for i, j in enumerate(monthlabels)
                               if j == month_num]
            items_dict[month_num] = _one_month_list

        ws = summary_wb.create_sheet(cat_name, num_of_ws)
        try:
            ws.column_dimensions["A"].width = max([len(i) for i in items]) + 2
        except:
            return summary_wb

        row = 0
        month = start_label[0]
        year = start_label[1]
        for number in range(1, n_of_months + 1):
            if number in monthlabels:
                row += 1
                ws.cell(row, 1).value = \
                    mdict[month] + " " + "20" + str(year) + " - " \
                    + str(round(sum(values_dict[number]), 2)) + "zł"
                ws.merge_cells(start_row=row, start_column=1, end_row=row,
                               end_column=2)
                ws.cell(row, 1).alignment = Alignment(horizontal="center")
                ws.cell(row, 1).fill = PatternFill(fgColor="93e1e6",
                                                   fill_type="solid")

                row += 1
                ws.cell(row, 1).value = "Co?"
                ws.cell(row, 2).value = "Kwota [zł]"
                ws.cell(row, 1).alignment = Alignment(horizontal="center")
                ws.cell(row, 2).alignment = Alignment(horizontal="center")
                ws.cell(row, 1).fill = PatternFill(fgColor="daf6f7",
                                                   fill_type="solid")
                ws.cell(row, 2).fill = PatternFill(fgColor="daf6f7",
                                                   fill_type="solid")

                for i, val in enumerate(values_dict[number]):
                    row += 1
                    ws.cell(row, 1).value = items_dict[number][i]
                    ws.cell(row, 2).value = val

            month += 1
            if month > 12:
                month -= 12
                year += 1

        for r in range(row):
            ws.cell(row=r + 1, column=1).border = thin_border
            ws.cell(row=r + 1, column=2).border = thin_border

        # 2a) Summary of the category (mean, median, std)
        # For monthly sums
        cat_spends_list = [sheet.cats_sums[cat_name]
                           for sheet in my_workbook.sheets_list]

        ws.merge_cells(start_row=1, start_column=7, end_row=1,
                       end_column=9)
        ws.cell(1, 7).value = "Dla sum miesięcznych:"
        ws.cell(2, 7).value = "Średnia [zł]:"
        ws.cell(3, 7).value = round(np.mean(cat_spends_list), 2)
        ws.column_dimensions["G"].width = 14
        ws.cell(2, 8).value = "Mediana [zł]:"
        ws.cell(3, 8).value = round(np.median(cat_spends_list), 2)
        ws.column_dimensions["H"].width = 14
        ws.cell(2, 9).value = "Std [zł]:"
        ws.cell(3, 9).value = round(np.std(cat_spends_list), 2)

        for r in range(1, 4):
            for c in range(7, 10):
                ws.cell(r, c).alignment = Alignment(horizontal="center")
                ws.cell(r, c).border = thin_border

        # For individual spendings
        all_values = [v for sublist in values_dict.values() for v in sublist]

        ws.merge_cells(start_row=7, start_column=7, end_row=7,
                       end_column=9)
        ws.cell(7, 7).value = "Dla wydatków indywidualnych:"
        ws.cell(8, 7).value = "Średnia [zł]:"
        ws.cell(9, 7).value = round(np.mean(all_values), 2)
        ws.cell(8, 8).value = "Mediana [zł]:"
        ws.cell(9, 8).value = round(np.median(all_values), 2)
        ws.cell(8, 9).value = "Std [zł]:"
        ws.cell(9, 9).value = round(np.std(all_values), 2)

        for r in range(7, 10):
            for c in range(7, 10):
                ws.cell(r, c).alignment = Alignment(horizontal="center")
                ws.cell(r, c).border = thin_border

        return summary_wb

    list_of_categories \
        = ["Rzeczy i sprzęty", "Hobby i przyjemności", "Transport i noclegi",
           "Podróże", "Abonamenty i usługi", "Leki i zdrowie",
           "Książki i nauka"]

    for i, cat in enumerate(list_of_categories):
        summary_wb = export_to_excel(cat, i + 1)

    summary_wb.save(results_dir + "/" + total_label + " - podsumowanie.xlsx")
