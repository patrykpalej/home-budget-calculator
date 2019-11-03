import json
import os
import datetime
import numpy as np
from classes import *
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill


# 1. Importing total data and config, choosing data in range
spendings_file_path = os.getcwd() + "/data/total_test.xlsx"

with open("spendings_finder/config.json", encoding='utf-8') as file:
    config_json = json.load(file)

config_list = config_json["keywords"]
period = config_json["period"]

start_date = datetime.date(year=int(config_json["period"][1]),
                           month=int(config_json["period"][0]), day=1)

end_date = datetime.date(year=int(config_json["period"][3]),
                         month=int(config_json["period"][2]), day=1)

all_sheetnames = MyWorkbook(spendings_file_path).mywb.sheetnames
sheetnames_in_range = []

for date_label in all_sheetnames:
    current_date = datetime.date(int(date_label[-2:]), int(date_label[:2]), 1)
    if (current_date >= start_date) and (current_date <= end_date):
        sheetnames_in_range.append(date_label)

myWorkbook = MyWorkbook(spendings_file_path, sheetnames_in_range)

spend_values = myWorkbook.spends_values_yr
spend_items = myWorkbook.spends_items_yr
spend_monthlabels = myWorkbook.spends_monthlabel_yr

month_year_label_dict = {}
for month in range(len(sheetnames_in_range)):
    month_int = start_date.month + month \
                    if start_date.month + month <= 12 \
                    else (start_date.month + month) % 12
    month_dict = {1: "Styczeń", 2: "Luty", 3: "Marzec", 4: "Kwiecień",
                  5: "Maj", 6: "Czerwiec", 7: "Lipiec", 8: "Sierpień",
                  9: "Wrzesień", 10: "Październik", 11: "Listopad",
                  12: "Grudzień"}

    year_label = " 20" + str(start_date.year
                             + int(format((np.floor(
                                    (start_date.month + month-1)/12)), '.0f')))

    month_year_label_dict[month+1] = month_dict[month_int] + year_label


# 2. Preparation of list of encoded dictionaries of spendings
spends_dicts_list = list()

for phrase in config_list:
    phrase_dict = {"cat_sums": {}, "date_spends": {}}

    for cat in spend_items.keys():
        for i, item in enumerate(spend_items[cat]):
            if phrase.lower() in item.lower():
                if cat in phrase_dict["cat_sums"].keys():
                    phrase_dict["cat_sums"][cat] += spend_values[cat][i]
                else:
                    phrase_dict["cat_sums"][cat] = spend_values[cat][i]

                if spend_monthlabels[cat][i] \
                        in phrase_dict["date_spends"].keys():
                    phrase_dict["date_spends"][spend_monthlabels[cat][i]]\
                        .append([spend_items[cat][i], spend_values[cat][i]])
                else:
                    phrase_dict["date_spends"][spend_monthlabels[cat][i]] \
                        = [[spend_items[cat][i], spend_values[cat][i]]]

    spends_dicts_list.append(phrase_dict)

# 3. Creating xlsx file for the results
output_xlsx = openpyxl.Workbook()
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

for p, phrase in enumerate(config_list):
    max_col_width = 0

    ws = output_xlsx[output_xlsx.sheetnames[-1]]
    ws.title = phrase
    phrase_dict = spends_dicts_list[p]

    # a) cat_sums table
    try:
        ws.column_dimensions["A"].width = max([len(cat) for cat in
                                               phrase_dict["cat_sums"].keys()])
    except ValueError:
        pass

    row = 1
    for cat in phrase_dict["cat_sums"].keys():
        ws.cell(row, 1).value = cat
        ws.cell(row, 2).value = phrase_dict["cat_sums"][cat]

        ws.cell(row, 1).border = thin_border
        ws.cell(row, 2).border = thin_border

        ws.cell(row, 1).fill = PatternFill(fgColor='daf6f7', fill_type='solid')

        row += 1

    # b) date_spends table
    row = 1
    for date in sorted(phrase_dict["date_spends"].keys()):
        # !!! convert number of month into {Month yyyy} format
        ws.merge_cells(None, row, 8, row, 9)
        ws.cell(row, 8).value = month_year_label_dict[date]
        ws.cell(row, 8).alignment = Alignment(horizontal='center')
        ws.cell(row, 8).fill = PatternFill(fgColor='93e1e6', fill_type='solid')
        ws.cell(row, 8).border = thin_border
        row += 1

        ws.cell(row, 8).value = "Co?"
        ws.cell(row, 9).value = "Kwota [zł]"
        ws.cell(row, 8).alignment = Alignment(horizontal='center')
        ws.cell(row, 9).alignment = Alignment(horizontal='center')
        ws.cell(row, 8).fill = PatternFill(fgColor='daf6f7', fill_type='solid')
        ws.cell(row, 9).fill = PatternFill(fgColor='daf6f7', fill_type='solid')
        ws.cell(row, 8).border = thin_border
        ws.cell(row, 9).border = thin_border

        row += 1
        for spend in phrase_dict["date_spends"][date]:
            ws.cell(row, 8).value = spend[0]
            ws.cell(row, 9).value = spend[1]
            ws.cell(row, 8).border = thin_border
            ws.cell(row, 9).border = thin_border

            max_col_width = max(max_col_width, len(spend[0]))
            row += 1

    try:
        ws.column_dimensions["H"].width = max_col_width
    except NameError:
        pass

    if len(config_list)-2 >= p:
        output_xlsx.create_sheet("name", p+1)

# 4. Postprocessing
results_dir = os.getcwd() + "/spendings_finder"
if not os.path.exists(results_dir):
    os.mkdir(results_dir)

output_xlsx.save(results_dir + "/Przedział czasu.xlsx")
